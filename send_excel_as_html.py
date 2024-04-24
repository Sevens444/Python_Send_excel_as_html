import os
import pandas as pd
import numpy as np
import openpyxl as opx
from itertools import chain
from win32com.client.gencache import EnsureDispatch
from win32com.client import constants
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import shutil

email_send_to = [
        'Reb@sibur.ru',
        'Dul@sibur.ru',
        'Pol@sibur.ru',
        'ObraztsovNK@sibur.ru'
        ]

_email_login = os.getenv('MAIL_USERNAME')
_email_password = os.getenv('MAIL_PASSWORD')

#%%
'''  Чтение файлов  '''
dir_reports = r"\\S-files\КАЧЕСТВО ПРОДУКЦИИ\Отчет ТМБ Дебалансы"
dir_mon = dir_reports + '\\' + [x for x in os.listdir(dir_reports) if x[:14] == 'ТМБ. Мономеры '][0]
dir_ras = dir_reports + '\\' + [x for x in os.listdir(dir_reports) if x[:14] == 'ТМБ. Растворит'][0]
dir_table_to_send = dir_reports + '\\' + 'Table debalans.xlsx'

df_mon = pd.read_excel(dir_mon, sheet_name='Динамика дебалансов', header=None)
df_ras_tep = pd.read_excel(dir_ras, sheet_name='Динамика дебалансов Т', header=None)
df_ras_skd = pd.read_excel(dir_ras, sheet_name='Динамика дебалансов С', header=None)
df_ras_dst = pd.read_excel(dir_ras, sheet_name='Динамика дебалансов Д', header=None)
df_ras_nod = pd.read_excel(dir_ras, sheet_name='Динамика дебалансов Н', header=None)

'''  Мономеры. Формирование таблицы  '''
for col_last in range(6,38):
    if df_mon.iat[4, col_last - 1] == 'Итого':
           break

df_mon_tabl = df_mon.iloc[5:38, 4:col_last]
df_mon_tabl.columns = df_mon.iloc[4, 4:col_last].to_list()
df_mon_tabl.index = df_mon.iloc[5:38, 3].to_list()

#%%
'''  Растворители. Формирование таблиц  '''
for col_itog in range(5,36):
    if np.isnan(df_ras_tep.iloc[8,col_itog]):
        col_itog += 1
        break

df_ras_tep_tabl = df_ras_tep.iloc[7:25, 4:col_itog]
df_ras_tep_tabl.columns = df_ras_tep.iloc[6, 4:col_itog].to_list()
df_ras_tep_tabl.index = df_ras_tep.iloc[7:25, 3].to_list()

df_ras_skd_tabl = df_ras_skd.iloc[7:21, 4:col_itog]
df_ras_skd_tabl.columns = df_ras_skd.iloc[6, 4:col_itog].to_list()
df_ras_skd_tabl.index = df_ras_skd.iloc[7:21, 3].to_list()

df_ras_dst_tabl = df_ras_dst.iloc[7:19, 4:col_itog]
df_ras_dst_tabl.columns = df_ras_dst.iloc[6, 4:col_itog].to_list()
df_ras_dst_tabl.index = df_ras_dst.iloc[7:19, 3].to_list()

df_ras_no_d_tabl = df_ras_nod.iloc[7:21, 4:col_itog]
df_ras_no_d_tabl.columns = df_ras_nod.iloc[6, 4:col_itog].to_list()
df_ras_no_d_tabl.index = df_ras_nod.iloc[7:21, 3].to_list()

'''  Обнуление nan  '''
for df in [df_ras_tep_tabl, df_ras_skd_tabl, df_ras_dst_tabl, df_ras_no_d_tabl]:
    for _row in range(df.shape[0] - 3):
        for _col in range(1, df.shape[1]):
            if np.isnan(df.iloc[_row,_col]):
                df.iloc[_row,_col] = 0
                
    '''  Расчет итогов  '''
    df['Итого'] = 0
    for _row in range(df.shape[0]):
        if df.iloc[_row,0] == 'т':
            df.iloc[_row,-1] = df.iloc[_row, 1:-1].sum()
        else:
            df.iloc[_row,-1] = np.nan

#%%
'''  Сохранение  '''
with pd.ExcelWriter(dir_table_to_send) as writer:
    df_mon_tabl.to_excel(writer)
    df_ras_tep_tabl.to_excel(writer, startrow=36)
    df_ras_skd_tabl.to_excel(writer, startrow=57)
    df_ras_dst_tabl.to_excel(writer, startrow=74)
    df_ras_no_d_tabl.to_excel(writer, startrow=89)

#%%
'''  Заливка расхождений  '''
# Желтая заливка ячейки:  текущее < следующее значение
# Красная заливка ячейки: текущее < следующее < последующее значение
# Красный шрифт ячейки:   текущее - следующее > 30

wb = opx.load_workbook(dir_table_to_send)
ws = wb[wb.sheetnames[0]]

max_col = len(tuple(ws.columns))
# Мономеры
for row in range(2, 28+1, 2):
    for col in range(4, col_last-2):
        if col == 4:
            val_cur, val_past_1 = ws.cell(row=row, column=col).value, ws.cell(row=row, column=col - 1).value
            if (val_cur > 0 and val_past_1 > 0) or (val_cur < 0 and val_past_1 < 0):
                if abs(val_cur) > abs(val_past_1):
                    ws.cell(row=row  , column=col).fill = opx.styles.fills.PatternFill(patternType='solid', fgColor='FFFF66')
                    ws.cell(row=row+1, column=col).fill = opx.styles.fills.PatternFill(patternType='solid', fgColor='FFFF66')
        else:
            val_cur, val_past_1, val_past_2 = ws.cell(row=row, column=col).value, ws.cell(row=row, column=col - 1).value, ws.cell(row=row, column=col - 2).value
            if (val_cur > 0 and val_past_1 > 0) or (val_cur < 0 and val_past_1 < 0):
                if abs(val_cur) > abs(val_past_1):
                    ws.cell(row=row  , column=col).fill = opx.styles.fills.PatternFill(patternType='solid', fgColor='FFFF66')
                    ws.cell(row=row+1, column=col).fill = opx.styles.fills.PatternFill(patternType='solid', fgColor='FFFF66')

            if (val_cur > 0 and val_past_1 > 0 and val_past_2 > 0) or (val_cur < 0 and val_past_1 < 0 and val_past_2 < 0):
                if abs(val_cur) > abs(val_past_1) > abs(val_past_2):
                    ws.cell(row=row  , column=col).fill = opx.styles.fills.PatternFill(patternType='solid', fgColor='FF6666')
                    ws.cell(row=row+1, column=col).fill = opx.styles.fills.PatternFill(patternType='solid', fgColor='FF6666')

        percent_cur, percent_past = ws.cell(row=row + 1, column=col).value, ws.cell(row=row + 1, column=col - 1).value
        if abs(percent_cur - percent_past) > 30:
            ws.cell(row=row,   column=col).font = opx.styles.Font(bold=True, color='A3051F')
            ws.cell(row=row+1, column=col).font = opx.styles.Font(bold=True, color='A3051F')
# Растворители
for row in chain(range(2, 28+1, 2), range(38, 48+1, 2), range(59, 65+1, 2), range(76, 80+1, 2), range(91, 97+1, 2)):
    for col in range(4, col_itog-4):
        if col == 4:
            val_cur, val_past_1 = ws.cell(row=row, column=col).value, ws.cell(row=row, column=col - 1).value
            if (val_cur > 0 and val_past_1 > 0) or (val_cur < 0 and val_past_1 < 0):
                if abs(val_cur) > abs(val_past_1):
                    ws.cell(row=row  , column=col).fill = opx.styles.fills.PatternFill(patternType='solid', fgColor='FFFF66')
                    ws.cell(row=row+1, column=col).fill = opx.styles.fills.PatternFill(patternType='solid', fgColor='FFFF66')
        else:
            val_cur, val_past_1, val_past_2 = ws.cell(row=row, column=col).value, ws.cell(row=row, column=col - 1).value, ws.cell(row=row, column=col - 2).value
            if (val_cur > 0 and val_past_1 > 0) or (val_cur < 0 and val_past_1 < 0):
                if abs(val_cur) > abs(val_past_1):
                    ws.cell(row=row  , column=col).fill = opx.styles.fills.PatternFill(patternType='solid', fgColor='FFFF66')
                    ws.cell(row=row+1, column=col).fill = opx.styles.fills.PatternFill(patternType='solid', fgColor='FFFF66')

            if (val_cur > 0 and val_past_1 > 0 and val_past_2 > 0) or (val_cur < 0 and val_past_1 < 0 and val_past_2 < 0):
                if abs(val_cur) > abs(val_past_1) > abs(val_past_2):
                    ws.cell(row=row  , column=col).fill = opx.styles.fills.PatternFill(patternType='solid', fgColor='FF6666')
                    ws.cell(row=row+1, column=col).fill = opx.styles.fills.PatternFill(patternType='solid', fgColor='FF6666')

        percent_cur, percent_past = ws.cell(row=row + 1, column=col).value, ws.cell(row=row + 1, column=col - 1).value
        if abs(percent_cur - percent_past) > 30:
            ws.cell(row=row,   column=col).font = opx.styles.Font(bold=True, color='A3051F')
            ws.cell(row=row+1, column=col).font = opx.styles.Font(bold=True, color='A3051F')

'''  Выравнивание, форматирование  '''
for row in range(2, 105):
    if row in [1,37,58, 75, 90]:
        continue
    ws.cell(row, 1).alignment = opx.styles.Alignment(horizontal='left', vertical='center')
    if row > 36:
        ws.cell(row,col_last-2).alignment = opx.styles.Alignment(horizontal='right',vertical='center')
    else:
        ws.cell(row,col_itog-3).alignment = opx.styles.Alignment(horizontal='right',vertical='center')

    for col in range(3, col_last-2):
        ws.cell(row, col).number_format = '# ##0.00'

for row in [1,37,58, 75, 90]:
    for col in range(3, max_col):
        ws.cell(row, col).number_format = 'd mmm'

ws.column_dimensions['A'].width = 28

wb.save(dir_table_to_send)

#%%
'''  Конвертирую всю excel книгу в html  '''
xl = EnsureDispatch('Excel.Application')
wb = xl.Workbooks.Open(dir_table_to_send)
wb.SaveAs("D:\\ObraztsovNK\\book.html", constants.xlHtml)
xl.Workbooks.Close()
xl.Quit()
del xl

#%%
'''  Формирование html таблицы  '''
table_html = open('D:\\ObraztsovNK\\book.files\\sheet001.html', 'r').read()
table_css = open('D:\\ObraztsovNK\\book.files\\stylesheet.css', 'r').read()

cont_tabl = '<html> <head> <style type="text/css">\n' + \
            table_css + '</style> </head>' + \
            table_html[table_html.find("<body link=blue vlink=purple>"):-18] + \
            '<p></font><a href="\\\\S-files\КАЧЕСТВО ПРОДУКЦИИ\Отчет ТМБ Дебалансы"><span style="color:blue">Ссылка на оригинал</span></a></p><p>' + \
            'Желтая заливка - при первом увеличении дебаланса в тоннах<br>\n' + \
            'Красная заливка - при последующем увеличении дебаланса в тоннах<br>\n' + \
            'Красный цвет текста - если текущий процент отклонения больше предыдущего на 30 единиц<br>\n' + \
            '</p></body></html>'

#%%
'''  Отправка сообщения  '''
msg = MIMEMultipart()
msg['Subject'] = 'Таблица дебалансов'
msg['From'] = 'rprt@sibur.ru'
msg['To'] = ", ".join(email_send_to)
msg.attach(MIMEText(cont_tabl, 'html'))

server = smtplib.SMTP('smtp.sibur.local', port=400)
server.ehlo()
server.starttls()
server.ehlo()
server.login(_email_login, _email_password)
server.sendmail(msg['From'], send_to, msg.as_string())
server.quit()

#%%
'''  Удаление вспомогательных файлов  '''
os.remove("D:\\ObraztsovNK\\book.html")
shutil.rmtree("D:\\ObraztsovNK\\book.files")
