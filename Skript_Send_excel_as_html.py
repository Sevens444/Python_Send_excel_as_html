import os
import pandas as pd
import numpy as np
import openpyxl as opx
from itertools import chain
from win32com.client.gencache import EnsureDispatch
from win32com.client import constants
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import shutil

email_send_to = [
        'Reb@vsk.sibur.ru',
        'Dul@vsk.sibur.ru',
        'Pol@vsk.sibur.ru',
        'ObraztsovNK@vsk.sibur.ru'
        ]

_MII_LOGIN = os.getenv('MAIL_USERNAME')
_MII_PASSWORD = os.getenv('MAIL_PASSWORD')

#%%
'''  Чтение файлов  '''
dir_reports = r"\\S-files\КАЧЕСТВО ПРОДУКЦИИ\Отчет ТМБ Дебалансы"
dir_Mon = dir_reports + '\\' + [x for x in os.listdir(dir_reports) if x[:14] == 'ТМБ. Мономеры '][0]
dir_Ras = dir_reports + '\\' + [x for x in os.listdir(dir_reports) if x[:14] == 'ТМБ. Растворит'][0]
dir_Tab_Deb_xl = dir_reports + '\\' + 'Table debalans.xlsx'

DS_Mon = pd.read_excel(dir_Mon, sheet_name='Динамика дебалансов', header=None)
DS_Ras_TEP = pd.read_excel(dir_Ras, sheet_name='Динамика дебалансов Т', header=None)
DS_Ras_SKD = pd.read_excel(dir_Ras, sheet_name='Динамика дебалансов С', header=None)
DS_Ras_DST = pd.read_excel(dir_Ras, sheet_name='Динамика дебалансов Д', header=None)
DS_Ras_NoD = pd.read_excel(dir_Ras, sheet_name='Динамика дебалансов Н', header=None)

'''  Мономеры. Формирование таблицы  '''
for col_last in range(6,38):
    if DS_Mon.iat[4,col_last-1] == 'Итого':
           break

DS_Mon_Tabl = DS_Mon.iloc[5:38, 4:col_last]
DS_Mon_Tabl.columns = DS_Mon.iloc[4, 4:col_last].to_list()
DS_Mon_Tabl.index = DS_Mon.iloc[5:38, 3].to_list()

#%%
'''  Растворители. Формирование таблиц  '''
for col_itog in range(5,36):
    if np.isnan(DS_Ras_TEP.iloc[8,col_itog]):
        col_itog += 1
        break

DS_Ras_TEP_Tabl = DS_Ras_TEP.iloc[7:25, 4:col_itog]
DS_Ras_TEP_Tabl.columns = DS_Ras_TEP.iloc[6, 4:col_itog].to_list()
DS_Ras_TEP_Tabl.index = DS_Ras_TEP.iloc[7:25, 3].to_list()

DS_Ras_SKD_Tabl = DS_Ras_SKD.iloc[7:21, 4:col_itog]
DS_Ras_SKD_Tabl.columns = DS_Ras_SKD.iloc[6, 4:col_itog].to_list()
DS_Ras_SKD_Tabl.index = DS_Ras_SKD.iloc[7:21, 3].to_list()

DS_Ras_DST_Tabl = DS_Ras_DST.iloc[7:19, 4:col_itog]
DS_Ras_DST_Tabl.columns = DS_Ras_DST.iloc[6, 4:col_itog].to_list()
DS_Ras_DST_Tabl.index = DS_Ras_DST.iloc[7:19, 3].to_list()

DS_Ras_NoD_Tabl = DS_Ras_NoD.iloc[7:21, 4:col_itog]
DS_Ras_NoD_Tabl.columns = DS_Ras_NoD.iloc[6, 4:col_itog].to_list()
DS_Ras_NoD_Tabl.index = DS_Ras_NoD.iloc[7:21, 3].to_list()

'''  Обнуление nan  '''
for _Tabl in [DS_Ras_TEP_Tabl, DS_Ras_SKD_Tabl, DS_Ras_DST_Tabl, DS_Ras_NoD_Tabl]:
    for _row in range(_Tabl.shape[0]-3):
        for _col in range(1,_Tabl.shape[1]):
            if np.isnan(_Tabl.iloc[_row,_col]):
                _Tabl.iloc[_row,_col] = 0
                
'''  Расчет итогов  '''
    _Tabl['Итого'] = 0
    for _row in range(_Tabl.shape[0]):
        if _Tabl.iloc[_row,0] == 'т':
            _Tabl.iloc[_row,-1] = _Tabl.iloc[_row,1:-1].sum()
        else:
            _Tabl.iloc[_row,-1] = np.nan

#%%
'''  Сохранение  '''
with pd.ExcelWriter(dir_Tab_Deb_xl) as writer:
    DS_Mon_Tabl.to_excel(writer)
    DS_Ras_TEP_Tabl.to_excel(writer, startrow=36)
    DS_Ras_SKD_Tabl.to_excel(writer, startrow=57)
    DS_Ras_DST_Tabl.to_excel(writer, startrow=74)
    DS_Ras_NoD_Tabl.to_excel(writer, startrow=89)

#%%
'''  Заливка расхождений  '''
# Желтая заливка ячейки:  текущее < следующее значение
# Красная заливка ячейки: текущее < следующее < последующее значение
# Красный шрифт ячейки:   текущее - следующее > 30

wb = opx.load_workbook(dir_Tab_Deb_xl)
ws = wb[wb.sheetnames[0]]

max_col = len(tuple(ws.columns))
# Мономеры
for row in range(2, 28+1, 2):
    for col in range(4, col_last-2):
        if col == 4:
            VT, VP1 = ws.cell(row=row, column=col).value, ws.cell(row=row, column=col-1).value
            if (VT > 0 and VP1 > 0) or (VT < 0 and VP1 < 0):
                if abs(VT) > abs(VP1):
                    ws.cell(row=row  , column=col).fill = opx.styles.fills.PatternFill(patternType='solid', fgColor='FFFF66')
                    ws.cell(row=row+1, column=col).fill = opx.styles.fills.PatternFill(patternType='solid', fgColor='FFFF66')
        else:
            VT, VP1, VP2 = ws.cell(row=row, column=col).value, ws.cell(row=row, column=col-1).value, ws.cell(row=row, column=col-2).value
            if (VT > 0 and VP1 > 0) or (VT < 0 and VP1 < 0):
                if abs(VT) > abs(VP1):
                    ws.cell(row=row  , column=col).fill = opx.styles.fills.PatternFill(patternType='solid', fgColor='FFFF66')
                    ws.cell(row=row+1, column=col).fill = opx.styles.fills.PatternFill(patternType='solid', fgColor='FFFF66')

            if (VT > 0 and VP1 > 0 and VP2 > 0) or (VT < 0 and VP1 < 0 and VP2 < 0):
                if abs(VT) > abs(VP1) > abs(VP2):
                    ws.cell(row=row  , column=col).fill = opx.styles.fills.PatternFill(patternType='solid', fgColor='FF6666')
                    ws.cell(row=row+1, column=col).fill = opx.styles.fills.PatternFill(patternType='solid', fgColor='FF6666')

        ProcTek, ProcPst = ws.cell(row=row+1, column=col).value, ws.cell(row=row+1, column=col-1).value
        if abs(ProcTek - ProcPst) > 30:
            ws.cell(row=row,   column=col).font = opx.styles.Font(bold=True, color='A3051F')
            ws.cell(row=row+1, column=col).font = opx.styles.Font(bold=True, color='A3051F')
# Растворители
for row in chain(range(2, 28+1, 2), range(38, 48+1, 2), range(59, 65+1, 2), range(76, 80+1, 2), range(91, 97+1, 2)):
    for col in range(4, col_itog-4):
        if col == 4:
            VT, VP1 = ws.cell(row=row, column=col).value, ws.cell(row=row, column=col-1).value
            if (VT > 0 and VP1 > 0) or (VT < 0 and VP1 < 0):
                if abs(VT) > abs(VP1):
                    ws.cell(row=row  , column=col).fill = opx.styles.fills.PatternFill(patternType='solid', fgColor='FFFF66')
                    ws.cell(row=row+1, column=col).fill = opx.styles.fills.PatternFill(patternType='solid', fgColor='FFFF66')
        else:
            VT, VP1, VP2 = ws.cell(row=row, column=col).value, ws.cell(row=row, column=col-1).value, ws.cell(row=row, column=col-2).value
            if (VT > 0 and VP1 > 0) or (VT < 0 and VP1 < 0):
                if abs(VT) > abs(VP1):
                    ws.cell(row=row  , column=col).fill = opx.styles.fills.PatternFill(patternType='solid', fgColor='FFFF66')
                    ws.cell(row=row+1, column=col).fill = opx.styles.fills.PatternFill(patternType='solid', fgColor='FFFF66')

            if (VT > 0 and VP1 > 0 and VP2 > 0) or (VT < 0 and VP1 < 0 and VP2 < 0):
                if abs(VT) > abs(VP1) > abs(VP2):
                    ws.cell(row=row  , column=col).fill = opx.styles.fills.PatternFill(patternType='solid', fgColor='FF6666')
                    ws.cell(row=row+1, column=col).fill = opx.styles.fills.PatternFill(patternType='solid', fgColor='FF6666')

        ProcTek, ProcPst = ws.cell(row=row+1, column=col).value, ws.cell(row=row+1, column=col-1).value
        if abs(ProcTek - ProcPst) > 30:
            ws.cell(row=row,   column=col).font = opx.styles.Font(bold=True, color='A3051F')
            ws.cell(row=row+1, column=col).font = opx.styles.Font(bold=True, color='A3051F')

'''  Выравнивание, форматирование  '''
for row in range(2, 105):
    if row in [1,37,58, 75, 90]:
        continue
    ws.cell(row, 1).alignment = opx.styles.Alignment(horizontal='left', vertical='center')
    if row > 36:
        ws.cell(row,col_last-2).alignment = opx.styles.Alignment(horizontal='right',vertical='center')
    else:
        ws.cell(row,col_itog-3).alignment = opx.styles.Alignment(horizontal='right',vertical='center')

    for col in range(3, col_last-2):
        ws.cell(row, col).number_format = '# ##0.00'

for row in [1,37,58, 75, 90]:
    for col in range(3, max_col):
        ws.cell(row, col).number_format = 'd mmm'

ws.column_dimensions['A'].width = 28

wb.save(dir_Tab_Deb_xl)

#%%
'''  Конвертирую всю excel книгу в html  '''
xl = EnsureDispatch('Excel.Application')
wb = xl.Workbooks.Open(dir_Tab_Deb_xl)
wb.SaveAs("D:\\ObraztsovNK\\book.html", constants.xlHtml)
xl.Workbooks.Close()
xl.Quit()
del xl

#%%
'''  Формирование html таблицы  '''
Tab_Deb_ht = open('D:\\ObraztsovNK\\book.files\\sheet001.html', 'r').read()
Tab_Deb_cs = open('D:\\ObraztsovNK\\book.files\\stylesheet.css','r').read()

cont_tabl = '<html> <head> <style type="text/css">\n' +\
            Tab_Deb_cs + '</style> </head>' +\
            Tab_Deb_ht[Tab_Deb_ht.find("<body link=blue vlink=purple>"):-18] +\
            '<p></font><a href="\\\\S-files\КАЧЕСТВО ПРОДУКЦИИ\Отчет ТМБ Дебалансы"><span style="color:blue">Ссылка на оригинал</span></a></p><p>' +\
            'Желтая заливка - при первом увеличении дебаланса в тоннах<br>\n' +\
            'Красная заливка - при последующем увеличении дебаланса в тоннах<br>\n' +\
            'Красный цвет текста - если текущий процент отклонения больше предыдущего на 30 единиц<br>\n' +\
            '</p></body></html>'

#%%
'''  Отправка сообщения  '''
msg = MIMEMultipart()
msg['Subject'] = 'Таблица дебалансов'
msg['From'] = 'rprt@vsk.sibur.ru'
msg['To'] = ", ".join(email_send_to)
msg.attach(MIMEText(cont_tabl, 'html'))

server = smtplib.SMTP('smtp.sibur.local', port=465)
server.ehlo()
server.starttls()
server.ehlo()
server.login(_MII_LOGIN, _MII_PASSWORD)
server.sendmail(msg['From'], send_to, msg.as_string())
server.quit()

#%%
'''  Удаление вспомогательных файлов  '''
os.remove("D:\\ObraztsovNK\\book.html")
shutil.rmtree("D:\\ObraztsovNK\\book.files")
